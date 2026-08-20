import os, re
from typing import Dict, List, Tuple

from PIL import Image as PILImage
from io import BytesIO
import csv





from ..main_unit import *


#------------------------------------------------------------
# 安全导入检查 -- 将导入语句修改为以下形式

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed, Excel-related nodes will not be available")

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    OpenpyxlImage = None
    print("Warning: openpyxl.drawing.image not available")

try:
    from openpyxl.utils import get_column_letter
except ImportError:
    get_column_letter = None
    print("Warning: openpyxl.utils.get_column_letter not available")


#------------------------------------------------------------


class excel_search_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "search_content": ("STRING", {"default": ""}),
                "search_mode": (["Precise_search", "Fuzzy_search"], {"default": "Precise_search"}),
            },
            "optional": {   } 
        }

    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("debug", "row", "col")
    FUNCTION = "search_data"
    CATEGORY = "Apt_Preset/prompt/excel"

    def IS_CHANGED(): return float("NaN")

    def search_data(self, excel_path, sheet_name, search_content, search_mode):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}", None, None)
            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}", None, None)
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            results = []
            found_row = None
            found_col = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    cell_value_str = str(cell_value)
                    if (search_mode == "Precise_search" and cell_value_str == search_content) or \
                        (search_mode == "Fuzzy_search" and search_content in cell_value_str):
                        results.append(f"{sheet_name}|{row}|{col}|{cell_value}")
                        found_row = row
                        found_col = col

            workbook.close()
            del workbook
            if not results:
                return ("No results found.", None, None)
            return ("\n".join(results), found_row, found_col)
        except Exception as e:
            return (f"Error: {str(e)}", None, None)


class excel_row_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "col_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "col_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {} 
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_row_diff"
    CATEGORY = "Apt_Preset/prompt/excel"
    DESCRIPTION = """
    - col_data=2: 统计第2列,从上到下连续非空单元格总数count1
    - col_finish=3: 统计第3列,从上到下连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls): 
        return float("NaN")

    def excel_row_diff(self, excel_path, sheet_name, col_data, col_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(col_index):
                if col_index == 0:  # 跳过无效列
                    return 0
                count = 0
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=col_index).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(col_data)
            count2 = count_cells(col_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")


class excel_column_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "row_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_column_diff"
    CATEGORY = "Apt_Preset/prompt/excel"
    DESCRIPTION = """
    - row_data=2: 统计第2行,从左到右连续非空单元格总数count1
    - row_finish=3: 统计第3行,从左到右连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls):
        return float("NaN")

    def excel_column_diff(self, excel_path, sheet_name, row_data, row_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(row_index):
                if row_index == 0:  # 跳过无效行
                    return 0
                count = 0
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_index, column=col).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(row_data)
            count2 = count_cells(row_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")






#------------------------------------------------------------建设中------------------------




class excel_qwen_artistic:   
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "qwen_Image")
    DEFAULT_FONT_DESC = "Arial 字体，经典西文字体，字形规整，笔画简洁"

    @staticmethod
    def load_excel_data(excel_path: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            data = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    key = str(row[0]).strip()
                    value = str(row[1]) if row[1] is not None else ""
                    data[key] = value
            workbook.close()
            return data
        except:
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        preset_path = os.path.join(cls.EXCEL_DIR, "模版预设.xlsx")
        medium_path = os.path.join(cls.EXCEL_DIR, "文字载体.xlsx")
        toon_path = os.path.join(cls.EXCEL_DIR, "整体氛围.xlsx")
        font_path = os.path.join(cls.EXCEL_DIR, "文字类型.xlsx")
        effect_path = os.path.join(cls.EXCEL_DIR, "文字效果.xlsx")
        array_path = os.path.join(cls.EXCEL_DIR, "文字排版.xlsx")
        
        cls.sum_preset_data = cls.load_excel_data(preset_path)
        cls.text_medium_data = cls.load_excel_data(medium_path)
        cls.sum_toon_data = cls.load_excel_data(toon_path)
        cls.text_font_data = cls.load_excel_data(font_path)
        cls.text_effect_data = cls.load_excel_data(effect_path)
        cls.text_array_data = cls.load_excel_data(array_path)
        
        if not cls.sum_preset_data:
            cls.sum_preset_data = {"默认预设": ""}
        if not cls.text_medium_data:
            cls.text_medium_data = {"默认载体": "图像上"}
        if not cls.sum_toon_data:
            cls.sum_toon_data = {"默认氛围": "常规氛围"}
        if not cls.text_font_data:
            cls.text_font_data = {"默认字体": cls.DEFAULT_FONT_DESC}
        if not cls.text_effect_data:
            cls.text_effect_data = {"默认效果": "标准效果"}
        if not cls.text_array_data:
            cls.text_array_data = {"默认排版": "常规排版"}
        
        return {
            "required": {
                "sum_preset": (list(cls.sum_preset_data.keys()), {"label": "模版预设"}),
                "text_medium": (list(cls.text_medium_data.keys()), {"label": "背景类型"}),
                "text_array": (list(cls.text_array_data.keys()), {"label": "字体样式"}),
                "text_font": (list(cls.text_font_data.keys()), {"label": "字体类型"}),
                "text_effect": (list(cls.text_effect_data.keys()), {"label": "字体效果"}),
                "sum_toon": (list(cls.sum_toon_data.keys()), {"label": "整体氛围"}),
                "text_cn": ("STRING", {"default": "", "multiline": False, "label": "中文文本"}),
                "text_en": ("STRING", {"default": "", "multiline": False, "label": "英文文本"}),
                "object": ("STRING", {"default": "", "multiline": False, "label": "元素文本"}),
                "target": ("STRING", {"default": "", "multiline": False, "label": "目标文本"}),
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模板"}),
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{text_medium}": medium_val,
    "{text_array}": array_val,
    "{text_font}": current_font,
    "{text_effect}": effect_val,
    "{sum_toon}": toon_val,
    "{text_cn}": text_cn,
    "{text_en}": text_en,
    "{object}": object,
    "{target}": target,
    "{text}": text_show

    - 举例：
    输入："文字的特点是 {text_font}，{sum_toon}" ，两个可替换内容，选择了 "宋体"，"蓝色氛围"
    输出结果："文字的特点是 宋体，蓝色氛围"

    """


    def execute(self, sum_preset, text_medium, text_array, text_font, text_effect, sum_toon, 
                text_cn="", text_en="", object="", target="", custom=""):
        text_show = text_cn.strip() if text_cn.strip() else text_en.strip() if text_en.strip() else "{text}"
        
        # 获取medium_val，确保有默认值"图像上"
        medium_val = self.text_medium_data.get(text_medium, "在图像上")
        if not medium_val.strip():  # 检查是否为空
            medium_val = "在图像上"
            
        font_val = self.text_font_data.get(text_font, "")
        array_val = self.text_array_data.get(text_array, "")
        effect_val = self.text_effect_data.get(text_effect, "")
        toon_val = self.sum_toon_data.get(sum_toon, "")  # 获取sum_toon对应的第二列内容
        current_font = font_val if font_val else self.DEFAULT_FONT_DESC
        
        # 检查是否有自定义模板输入，如果有且非空则使用自定义模板，否则使用预设模板
        if custom and custom.strip():
            preset_val = custom.strip()
            sum_preset_empty = False  # 自定义模板视为非空预设
        else:
            preset_val = self.sum_preset_data.get(sum_preset, "").strip()
            sum_preset_empty = (not preset_val)
        
        if sum_preset_empty:
            # 构建核心部分
            core_base = f"{medium_val}，以{current_font}呈现出文字\"{text_show}\""
            additional_parts = []
            
            # 添加排版和效果
            if array_val:
                additional_parts.append(f"排版方式为{array_val}")
            if effect_val:
                additional_parts.append(f"字体效果为{effect_val}")
            
            # 确保sum_toon内容被添加到最后
            if toon_val:
                additional_parts.append(f"整体的{toon_val}")
            
            # 组合结果
            if additional_parts:
                result = f"{core_base}，{','.join(additional_parts)}"
            else:
                result = core_base
                
            return (result,)
        
        # 处理有预设的情况（包括自定义模板）
        replacements = {
            "{text_medium}": medium_val,
            "{text_array}": array_val,
            "{text_font}": current_font,
            "{text_effect}": effect_val,
            "{sum_toon}": toon_val,
            "{text_cn}": text_cn,
            "{text_en}": text_en,
            "{object}": object,
            "{target}": target,
            "{text}": text_show
        }
        
        preset_content = preset_val
        for placeholder, value in replacements.items():
            if value:
                preset_content = self.single_replace(preset_content, placeholder, value)
        
        # 已移除：如果预设中没有包含sum_toon，不再自动添加到最后
        
        return (preset_content,)




class excel_VedioPrompt:   
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_PATH = os.path.join(BASE_DIR, "video", "视频提示词.xlsx")

    @staticmethod
    def load_excel_sheet(excel_path: str, sheet_name: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {}
                
            sheet = workbook[sheet_name]
            data = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                    
                if row[0] is None:
                    continue
                    
                key = str(row[0]).strip()
                if key.startswith('#'):
                    continue
                    
                chinese_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                english_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                
                # 只在键是默认模板时才设置默认值
                if key == "默认模板" and not chinese_val and not english_val:
                    chinese_val = ""
                    english_val = ""
                    
                data[key] = {
                    'chinese': chinese_val,
                    'english': english_val
                }
            
            workbook.close()
            return data
        except Exception as e:
            print(f"加载表单 {sheet_name} 时出错: {e}")
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载所有数据
        cls.template_data = cls.load_excel_sheet(cls.EXCEL_PATH, "模板")
        cls.style_data = cls.load_excel_sheet(cls.EXCEL_PATH, "风格")
        cls.scene_data = cls.load_excel_sheet(cls.EXCEL_PATH, "场景")
        cls.light_data = cls.load_excel_sheet(cls.EXCEL_PATH, "环境光")
        cls.camera_data = cls.load_excel_sheet(cls.EXCEL_PATH, "镜头")
        cls.atmosphere_data = cls.load_excel_sheet(cls.EXCEL_PATH, "氛围")
        cls.solar_term_data = cls.load_excel_sheet(cls.EXCEL_PATH, "节气")
        
        # 为各数据设置默认值
        if not cls.style_data:
            cls.style_data = {"默认风格": {'chinese': "写实风格", 'english': "realistic style"}}
        if not cls.light_data:
            cls.light_data = {"默认光线": {'chinese': "自然光", 'english': "natural light"}}
        if not cls.scene_data:
            cls.scene_data = {"默认场景": {'chinese': "室内场景", 'english': "indoor scene"}}
        if not cls.camera_data:
            cls.camera_data = {"默认镜头": {'chinese': "标准镜头", 'english': "standard lens"}}
        if not cls.atmosphere_data:
            cls.atmosphere_data = {"默认氛围": {'chinese': "舒适氛围", 'english': "comfortable atmosphere"}}
        if not cls.solar_term_data:
            cls.solar_term_data = {"默认节气": {'chinese': "春季", 'english': "spring"}}
        
        # 仅在模板数据为空时才添加默认模板选项
        if not cls.template_data:
            cls.template_data = {"默认模板": {'chinese': "", 'english': ""}}
        
        return {
            "required": {
                "template": (list(cls.template_data.keys()), {"label": "模板"}),
                "roles":  ("STRING", {"default": "", "multiline": False, }),
                "style": (list(cls.style_data.keys()), {"label": "风格"}),
                "light": (list(cls.light_data.keys()), {"label": "环境光"}),
                "scene": (list(cls.scene_data.keys()), {"label": "场景"}),
                "camera": (list(cls.camera_data.keys()), {"label": "镜头"}),
                "atmosphere": (list(cls.atmosphere_data.keys()), {"label": "氛围"}),
                "solar_term": (list(cls.solar_term_data.keys()), {"label": "节气"}),
                "language": ("BOOLEAN", {"default": True, "label_on": "中文", "label_off": "英文"}),
                "object": ("STRING", {"default": "", "multiline": False, "label": "元素文本"}),
                "target": ("STRING", {"default": "", "multiline": False, "label": "目标文本"}),
 
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模版"})
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{roles}": roles_val,
    "{style}": style_val,
    "{light}": light_val,
    "{scene}": scene_val,
    "{camera}": camera_val,
    "{atmosphere}": atmosphere_val,
    "{solar_term}": solar_term_val,
    "{object}": object,
    "{target}": target

    - 举例：
    输入："{light}，女孩在{scene}玩耍" ，两个可替换内容，选择了 "白天"，"草坪"
    输出结果："白天，女孩在草坪玩耍"

    """


    def execute(self, style, light, scene, camera, atmosphere, solar_term, template,
                language=True, object="", target="", custom="", roles=""):
        if not roles:
            roles = ""
            
        lang = 'chinese' if language else 'english'
     
        style_val = self.style_data.get(style, {}).get(lang, "默认风格" if language else "default style")
        light_val = self.light_data.get(light, {}).get(lang, "自然光" if language else "natural light")
        scene_val = self.scene_data.get(scene, {}).get(lang, "室内场景" if language else "indoor scene")
        camera_val = self.camera_data.get(camera, {}).get(lang, "标准镜头" if language else "standard lens")
        atmosphere_val = self.atmosphere_data.get(atmosphere, {}).get(lang, "舒适氛围" if language else "comfortable atmosphere")
        solar_term_val = self.solar_term_data.get(solar_term, {}).get(lang, "春季" if language else "spring")
        roles_val = roles.strip()  # 获取角色输入值
        custom_val = custom.strip()  # 获取自定义内容
        
        # 检查是否使用自定义模板
        if custom_val and not (self.template_data.get(template, {}).get(lang, "").strip()):
            # 当模板为空且custom有内容时，先构建默认公式再追加custom内容
            parts = []
            if roles_val:
                parts.append(roles_val)
            parts.extend([style_val, light_val, scene_val, camera_val, atmosphere_val, solar_term_val])
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            
            # 追加custom内容
            if result:  # 如果已有内容，先加分隔符
                result += separator + custom_val
            else:  # 如果没有其他内容，直接使用custom内容
                result = custom_val
                
            return (result,)
        elif custom_val:
            # 当模板不为空但有custom内容时，直接使用custom内容
            template_val = custom_val
            use_formula = False
        else:
            # 获取选中的模板值
            template_val = self.template_data.get(template, {}).get(lang, "").strip()
            # 如果模板值为空，则使用公式
            use_formula = (not template_val)
        
        # 如果需要使用公式
        if use_formula:
            # 构建部分列表，包含角色（如果有输入）
            parts = []
            if roles_val:
                parts.append(roles_val)
            parts.extend([style_val, light_val, scene_val, camera_val, atmosphere_val, solar_term_val])
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            return (result,)
        
        # 否则使用模板替换
        replacements = {
            "{roles}": roles_val,
            "{style}": style_val,
            "{light}": light_val,
            "{scene}": scene_val,
            "{camera}": camera_val,
            "{atmosphere}": atmosphere_val,
            "{solar_term}": solar_term_val,
            "{object}": object,
            "{target}": target
        }
        
        template_content = template_val
        for placeholder, value in replacements.items():
            if value:
                template_content = self.single_replace(template_content, placeholder, value)
        
        return (template_content,)




class excel_roles:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_PATH = os.path.join(BASE_DIR, "video", "角色.xlsx")

    @staticmethod
    def load_excel_sheet(excel_path: str, sheet_name: str) -> dict:
        if not os.path.exists(excel_path):
            return {}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {}
                
            sheet = workbook[sheet_name]
            data = {}
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue
                    
                if row[0] is None:
                    continue
                    
                key = str(row[0]).strip()
                if key.startswith('#'):
                    continue
                    
                chinese_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                english_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                
                # 只在键是默认模板时才设置默认值
                if key == "默认模板" and not chinese_val and not english_val:
                    # 默认模板，将所有的文本用逗号隔开，串联起来
                    chinese_val = "模版,角色,服饰,发型,皮肤,脸型,耳朵,眼睛,眉毛,鼻子,嘴巴,身材,胸部,动作,表情"
                    english_val = "template,role,clothing,hairstyle,skin,face shape,ears,eyes,eyebrows,nose,mouth,figure,chest,action,expression"
                    
                data[key] = {
                    'chinese': chinese_val,
                    'english': english_val
                }
            
            workbook.close()
            return data
        except Exception as e:
            print(f"加载表单 {sheet_name} 时出错: {e}")
            return {}

    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        return re.sub(re.escape(target_clean), replacement_clean, text)

    @classmethod
    def INPUT_TYPES(cls):
        # 加载所有数据
        cls.template_data = cls.load_excel_sheet(cls.EXCEL_PATH, "模版")
        cls.role_data = cls.load_excel_sheet(cls.EXCEL_PATH, "角色")
        cls.clothing_data = cls.load_excel_sheet(cls.EXCEL_PATH, "服饰")
        cls.hairstyle_data = cls.load_excel_sheet(cls.EXCEL_PATH, "发型")
        cls.skin_data = cls.load_excel_sheet(cls.EXCEL_PATH, "皮肤")
        cls.face_shape_data = cls.load_excel_sheet(cls.EXCEL_PATH, "脸型")
        cls.ears_data = cls.load_excel_sheet(cls.EXCEL_PATH, "耳朵")
        cls.eyes_data = cls.load_excel_sheet(cls.EXCEL_PATH, "眼睛")
        cls.eyebrows_data = cls.load_excel_sheet(cls.EXCEL_PATH, "眉毛")
        cls.nose_data = cls.load_excel_sheet(cls.EXCEL_PATH, "鼻子")
        cls.mouth_data = cls.load_excel_sheet(cls.EXCEL_PATH, "嘴巴")
        cls.figure_data = cls.load_excel_sheet(cls.EXCEL_PATH, "身材")
        cls.chest_data = cls.load_excel_sheet(cls.EXCEL_PATH, "胸部")
        cls.action_data = cls.load_excel_sheet(cls.EXCEL_PATH, "动作")
        cls.expression_data = cls.load_excel_sheet(cls.EXCEL_PATH, "表情")
        
        # 为各数据设置默认值
        if not cls.role_data:
            cls.role_data = {"默认角色": {'chinese': "普通人", 'english': "ordinary person"}}
        if not cls.clothing_data:
            cls.clothing_data = {"默认服饰": {'chinese': "日常服装", 'english': "everyday clothing"}}
        if not cls.hairstyle_data:
            cls.hairstyle_data = {"默认发型": {'chinese': "短发", 'english': "short hair"}}
        if not cls.skin_data:
            cls.skin_data = {"默认皮肤": {'chinese': "自然肤色", 'english': "natural skin tone"}}
        if not cls.face_shape_data:
            cls.face_shape_data = {"默认脸型": {'chinese': "圆形脸", 'english': "round face"}}
        if not cls.ears_data:
            cls.ears_data = {"默认耳朵": {'chinese': "普通耳朵", 'english': "normal ears"}}
        if not cls.eyes_data:
            cls.eyes_data = {"默认眼睛": {'chinese': "黑色眼睛", 'english': "black eyes"}}
        if not cls.eyebrows_data:
            cls.eyebrows_data = {"默认眉毛": {'chinese': "自然眉形", 'english': "natural eyebrows"}}
        if not cls.nose_data:
            cls.nose_data = {"默认鼻子": {'chinese': "中等鼻子", 'english': "average nose"}}
        if not cls.mouth_data:
            cls.mouth_data = {"默认嘴巴": {'chinese': "自然嘴唇", 'english': "natural lips"}}
        if not cls.figure_data:
            cls.figure_data = {"默认身材": {'chinese': "匀称身材", 'english': "well-proportioned figure"}}
        if not cls.chest_data:
            cls.chest_data = {"默认胸部": {'chinese': "适中胸部", 'english': "moderate chest"}}
        if not cls.action_data:
            cls.action_data = {"默认动作": {'chinese': "站立", 'english': "standing"}}
        if not cls.expression_data:
            cls.expression_data = {"默认表情": {'chinese': "自然表情", 'english': "natural expression"}}
        
        # 仅在模板数据为空时才添加默认模板选项
        if not cls.template_data:
            cls.template_data = {"默认模板": {'chinese': "", 'english': ""}}
        
        return {
            "required": {
                "template": (list(cls.template_data.keys()), {"label": "模版"}),
                "role": (list(cls.role_data.keys()), {"label": "角色"}),
                "clothing": (list(cls.clothing_data.keys()), {"label": "服饰"}),
                "hairstyle": (list(cls.hairstyle_data.keys()), {"label": "发型"}),
                "skin": (list(cls.skin_data.keys()), {"label": "皮肤"}),
                "face_shape": (list(cls.face_shape_data.keys()), {"label": "脸型"}),
                "ears": (list(cls.ears_data.keys()), {"label": "耳朵"}),
                "eyes": (list(cls.eyes_data.keys()), {"label": "眼睛"}),
                "eyebrows": (list(cls.eyebrows_data.keys()), {"label": "眉毛"}),
                "nose": (list(cls.nose_data.keys()), {"label": "鼻子"}),
                "mouth": (list(cls.mouth_data.keys()), {"label": "嘴巴"}),
                "figure": (list(cls.figure_data.keys()), {"label": "身材"}),
                "chest": (list(cls.chest_data.keys()), {"label": "胸部"}),
                "action": (list(cls.action_data.keys()), {"label": "动作"}),
                "expression": (list(cls.expression_data.keys()), {"label": "表情"}),
                "language": ("BOOLEAN", {"default": True, "label_on": "中文", "label_off": "英文"})
            },
            "optional": {
                "custom": ("STRING", {"default": "", "multiline": True, "placeholder": "输入=编辑新模版"})
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("pos",)
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - 编辑新模板，重新排版，可替换代码：
    "{role}": "角色",
    "{clothing}": "服饰",
    "{hairstyle}": "发型",
    "{skin}": "肤色",
    "{faceshape}": "脸型",
    "{ears}": "耳朵",
    "{eyes}": "眼睛",
    "{eyebrows}": "眉毛",
    "{nose}": "鼻子",
    "{mouth}": "嘴巴",
    "{figure}": "身材",
    "{chest}": "胸部",
    "{action}": "动作",
    "{expression}": "表情"
    例如：
    "{role}显得很{expression}"   输出：警察显得很惊讶
    """


    def execute(self, template, role, clothing, hairstyle, skin, face_shape, ears, eyes,
                eyebrows, nose, mouth, figure, chest, action, expression, language=True,
                custom=""):
        lang = 'chinese' if language else 'english'
     
        template_val = self.template_data.get(template, {}).get(lang, "")
        role_val = self.role_data.get(role, {}).get(lang, "普通人" if language else "ordinary person")
        clothing_val = self.clothing_data.get(clothing, {}).get(lang, "日常服装" if language else "everyday clothing")
        hairstyle_val = self.hairstyle_data.get(hairstyle, {}).get(lang, "短发" if language else "short hair")
        skin_val = self.skin_data.get(skin, {}).get(lang, "自然肤色" if language else "natural skin tone")
        face_shape_val = self.face_shape_data.get(face_shape, {}).get(lang, "圆形脸" if language else "round face")
        ears_val = self.ears_data.get(ears, {}).get(lang, "普通耳朵" if language else "normal ears")
        eyes_val = self.eyes_data.get(eyes, {}).get(lang, "黑色眼睛" if language else "black eyes")
        eyebrows_val = self.eyebrows_data.get(eyebrows, {}).get(lang, "自然眉形" if language else "natural eyebrows")
        nose_val = self.nose_data.get(nose, {}).get(lang, "中等鼻子" if language else "average nose")
        mouth_val = self.mouth_data.get(mouth, {}).get(lang, "自然嘴唇" if language else "natural lips")
        figure_val = self.figure_data.get(figure, {}).get(lang, "匀称身材" if language else "well-proportioned figure")
        chest_val = self.chest_data.get(chest, {}).get(lang, "适中胸部" if language else "moderate chest")
        action_val = self.action_data.get(action, {}).get(lang, "站立" if language else "standing")
        expression_val = self.expression_data.get(expression, {}).get(lang, "自然表情" if language else "natural expression")
        
        custom_val = custom.strip()  # 获取自定义内容
        
        # 检查是否使用自定义模板
        if custom_val and not (self.template_data.get(template, {}).get(lang, "").strip()):
            # 当模板为空且custom有内容时，先构建默认公式再追加custom内容
            parts = [
                role_val, clothing_val, hairstyle_val, skin_val, face_shape_val,
                ears_val, eyes_val, eyebrows_val, nose_val, mouth_val,
                figure_val, chest_val, action_val, expression_val
            ]
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            
            # 追加custom内容
            if result:  # 如果已有内容，先加分隔符
                result += separator + custom_val
            else:  # 如果没有其他内容，直接使用custom内容
                result = custom_val
                
            return (result,)
        elif custom_val:
            # 当模板不为空但有custom内容时，直接使用custom内容
            template_val = custom_val
            use_formula = False
        else:
            # 获取选中的模板值
            template_val = self.template_data.get(template, {}).get(lang, "").strip()
            # 如果模板值为空，则使用公式
            use_formula = (not template_val)
        
        # 如果需要使用公式
        if use_formula:
            # 构建部分列表
            parts = [
                role_val, clothing_val, hairstyle_val, skin_val, face_shape_val,
                ears_val, eyes_val, eyebrows_val, nose_val, mouth_val,
                figure_val, chest_val, action_val, expression_val
            ]
            
            # 过滤空值
            parts = [part for part in parts if part]
            
            separator = "，" if language else ", "
            result = separator.join(parts)
            return (result,)

        replacements = {
            "{role}": role_val,
            "{clothing}": clothing_val,
            "{hairstyle}": hairstyle_val,
            "{skin}": skin_val,
            "{faceshape}": face_shape_val,
            "{ears}": ears_val,
            "{eyes}": eyes_val,
            "{eyebrows}": eyebrows_val,
            "{nose}": nose_val,
            "{mouth}": mouth_val,
            "{figure}": figure_val,
            "{chest}": chest_val,
            "{action}": action_val,
            "{expression}": expression_val
        }
        
        template_content = template_val
        for placeholder, value in replacements.items():
            if value:
                template_content = self.single_replace(template_content, placeholder, value)
        
        if not template_content.strip():
            template_content = "None"
        return (template_content,)




class excel_insert_image_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
                "img_height": ("INT", {"default": 256, "max": 2048, "min": 64}),
                "image": ("IMAGE",),
            },
            "optional": {
                "debug_output": ("BOOLEAN", {"default": False, "label_on": "启用", "label_off": "禁用"}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_image"
    CATEGORY = "Apt_Preset/prompt/excel"
    OUTPUT_NODE = True  # 标记为输出节点

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_image(self, excel_path, sheet_name, row, column, img_height, image, debug_output=False):
        try:
            target_row = max(1, row)
            target_col = max(1, column)

            if not os.path.exists(excel_path):
                error_msg = f"Error: Excel file does not exist at path: {excel_path}"
                if debug_output:
                    print(error_msg)
                return (error_msg,)
            if not os.access(excel_path, os.W_OK):
                error_msg = f"Error: No write permission for Excel file at path: {excel_path}"
                if debug_output:
                    print(error_msg)
                return (error_msg,)

            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            cell_address = get_column_letter(target_col) + str(target_row)

            if hasattr(image, 'cpu'):
                image_np = image.cpu().numpy()
            else:
                image_np = image

            if len(image_np.shape) == 4:
                image_np = image_np[0]

            if image_np.max() <= 1.0:
                image_np = (image_np * 255).astype('uint8')
            else:
                image_np = image_np.astype('uint8')

            pil_img = PILImage.fromarray(image_np)

            width, height = pil_img.size
            scale = img_height / height
            target_width = int(width * scale)
            resized_img = pil_img.resize((target_width, img_height), PILImage.LANCZOS)

            img_byte_arr = BytesIO()
            resized_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            openpyxl_img = OpenpyxlImage(img_byte_arr)

            column_letter = get_column_letter(target_col)
            
            sheet.column_dimensions[column_letter].width = target_width * 0.75 / 7
            
            sheet.row_dimensions[target_row].height = 192

            sheet.add_image(openpyxl_img, cell_address)

            workbook.save(excel_path)
            workbook.close()

            success_msg = "Image inserted and scaled successfully!"
            if debug_output:
                print(success_msg)
            return (success_msg,)

        except PermissionError as pe:
            error_msg = f"Permission Error: {str(pe)}"
            if debug_output:
                print(error_msg)
            return (error_msg,)
        except Exception as e:
            error_msg = f"Error: {str(e)}"
            if debug_output:
                print(error_msg)
            return (error_msg,)




class excel_read_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("data",)
    FUNCTION = "excel_read"
    CATEGORY = "Apt_Preset/prompt/excel"

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def excel_read(self, excel_path, sheet_name, row, column):
        try:
            target_row = max(1, row)
            target_col = max(1, column)

            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)

            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}",)

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            cell_value = sheet.cell(row=target_row, column=target_col).value
            result = str(cell_value) if cell_value is not None else ""

            workbook.close()
            del workbook

            return (result,)

        except Exception as e:
            return (f"Error: {str(e)}",)



class excel_write_data_easy:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row": ("INT", {"default": 1, "min": 1}),
                "column": ("INT", {"default": 1, "min": 1}),
                "data": ("STRING", {"forceInput": True, }),
            },
            "optional": {
            }
        }
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_data"
    CATEGORY = "Apt_Preset/prompt/excel"
    OUTPUT_NODE = True
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")
    def write_data(self, excel_path, sheet_name, row, column, data,):

        try:
            target_row = max(1, row)
            target_col = max(1, column)
            if not os.path.exists(excel_path):
                error_msg = f"Error: File does not exist at path: {excel_path}"

            if not os.access(excel_path, os.W_OK):
                error_msg = f"Error: No write permission for file at path: {excel_path}"

            workbook = openpyxl.load_workbook(excel_path)
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            sheet.cell(row=target_row, column=target_col).value = None
            if data.strip():
                sheet.cell(row=target_row, column=target_col).value = data.strip()
                sheet.cell(row=target_row, column=target_col).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')

                success_msg = f"成功将数据整体写入行 {target_row} 列 {target_col}"
            else:
                success_msg = "未写入数据，输入内容为空"
            workbook.save(excel_path)
            workbook.close()

            return (success_msg,)
        except PermissionError as pe:
            error_msg = f"权限错误: {str(pe)}"

            return (error_msg,)
        except Exception as e:
            error_msg = f"错误: {str(e)}"
            return (error_msg,)
    



class excel_Prompter:  
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "global")
    CONFIG_FILE = os.path.join(EXCEL_DIR, "config.xlsx")
    excel_data: Dict[str, Dict] = {}
    excel_files: List[str] = []
    
    @classmethod
    def load_config(cls) -> List[str]:
        if not os.path.exists(cls.CONFIG_FILE):
            print(f"配置文件不存在: {cls.CONFIG_FILE}")
            return []
        try:
            workbook = openpyxl.load_workbook(cls.CONFIG_FILE, read_only=True, data_only=True)
            sheet = workbook.active
            xlsx_files = []
            row = 1
            while True:
                cell_value = sheet.cell(row=row, column=2).value
                if cell_value is None:
                    break
                file_name = str(cell_value).strip()
                if file_name and file_name.lower().endswith('.xlsx'):
                    xlsx_files.append(file_name)
                row += 1
            workbook.close()
            cls.excel_files = [os.path.splitext(f)[0] for f in xlsx_files]
            return xlsx_files
        except Exception as e:
            print(f"读取配置文件失败: {e}")
            return []
    
    @classmethod
    def load_all_excels(cls) -> None:
        xlsx_files = cls.load_config()
        cls.excel_data = {}
        for file in xlsx_files:
            key = os.path.splitext(file)[0]
            file_path = os.path.join(cls.EXCEL_DIR, file)
            cls.excel_data[key] = cls.load_excel(file_path)
    
    @staticmethod
    def load_excel(excel_path: str) -> dict:
        if not os.path.exists(excel_path):
            print(f"Excel文件不存在: {excel_path}")
            return {"文件不存在": ["", "", ""]}
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            headers = next(sheet.iter_rows(values_only=True))
            id_col = headers.index('ID') if 'ID' in headers else 0
            pos_col = headers.index('Positive') if 'Positive' in headers else 1
            neg_col = headers.index('Negative') if 'Negative' in headers else 2
            help_col = headers.index('Help') if 'Help' in headers else 3
            data = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[id_col]:
                    pos_val = str(row[pos_col] or '') if len(row) > pos_col else ''
                    neg_val = str(row[neg_col] or '') if len(row) > neg_col else ''
                    help_val = str(row[help_col] or '') if len(row) > help_col else ''
                    data[row[id_col]] = [pos_val, neg_val, help_val]
            workbook.close()
            return data if data else {"无有效数据": ["", "", ""]}
        except Exception as e:
            print(f"加载Excel失败: {excel_path} - {e}")
            return {"加载失败": ["", "", ""]}
    
    @staticmethod
    def split_with_quotes(s):
        pattern = r'"([^"]*)"|\s*([^,]+)'
        matches = re.finditer(pattern, s)
        return [match.group(1) or match.group(2).strip() for match in matches if match.group(1) or match.group(2).strip()]
    
    def single_replace(self, text, target, replacement):
        if not target or not replacement:
            return text
        target_clean = target.strip('"').strip()
        replacement_clean = replacement.strip('"').strip()
        pattern = re.escape(target_clean)
        return re.sub(pattern, replacement_clean, text)
    
    def multi_replace(self, text, multi_targets, multi_replacements):
        if not multi_targets or not multi_replacements:
            return text
        targets = multi_targets.split('@')
        replacements = multi_replacements.split('@')
        min_len = min(len(targets), len(replacements))
        targets = targets[:min_len]
        replacements = replacements[:min_len]
        result = text
        for target, replacement in zip(targets, replacements):
            result = self.single_replace(result, target.strip(), replacement.strip())
        return result

    @classmethod
    def INPUT_TYPES(cls):
        cls.load_all_excels()
        input_config = {
            "required": {

            },
            "optional": {
                "language": ("BOOLEAN", {"default": True, "label_on": "英文", "label_off": "中文"}),
            }
        }
        for file_key in cls.excel_files:
            options = list(cls.excel_data[file_key].keys())
            options.insert(0, "None")
            input_config["required"][file_key] = (options,)
        return input_config
    
    RETURN_TYPES = ("STRING", )  
    RETURN_NAMES = ("pos", )  
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    def execute(self, language=True, **kwargs):
        excel_positives = []
        excel_negatives = []
        
        for file_key in self.excel_files:
            selected_value = kwargs.get(file_key)
            if selected_value == "None":
                continue
            if selected_value in self.excel_data[file_key]:
                excel_data = self.excel_data[file_key][selected_value]
                pos_parts = [part.strip() for part in self.split_with_quotes(excel_data[0]) if part.strip()]
                neg_parts = [part.strip() for part in self.split_with_quotes(excel_data[1]) if part.strip()]
                
                if pos_parts:
                    selected_pos = pos_parts[0]
                    excel_positives.append(selected_pos)
                if neg_parts:
                    selected_neg = neg_parts[0]
                    excel_negatives.append(selected_neg)
        
        positive_prompt = ", ".join(filter(None, excel_positives))
        negative_prompt = ", ".join(filter(None, excel_negatives))
        
        final_output = negative_prompt if language else positive_prompt
        
        if not final_output:
            final_output = "❌ No valid excel content"
        
        return (final_output, )











